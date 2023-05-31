import openpyxl
# File -->Workbook --->Sheets ---> Rows ----> Cells

file="D:\RESHMA PEDDETI\Skills.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["India"]


rows=sheet.max_row #count no. of rows in excel
columns=sheet.max_column #count no. of columns in excel

for r in range(1,rows+1):
    for c in range(1,columns+1):
        print(sheet.cell(r,c).value,end='    ')
    print()
