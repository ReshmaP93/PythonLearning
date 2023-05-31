import openpyxl
# File -->Workbook --->Sheets ---> Rows ----> Cells

#file="D:\RESHMA PEDDETI\TestDataforDataDriveninPython.xlsx"
#workbook=openpyxl.load_workbook(file)
#sheet=workbook.active #get active sheet from excel
#for r in range(1,6):
    #for c in range(1,4):
        #sheet.cell(r,c).value='Welcome'
#workbook.save(file)

#Multiple Data
file="D:\RESHMA PEDDETI\TestDataforDataDriveninPython.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook.active #get active sheet from excel

sheet.cell(1,1).value=123
sheet.cell(1,2).value="Smith"
sheet.cell(1,3).value="Engineer"

sheet.cell(2,1).value=456
sheet.cell(2,2).value="Callie"
sheet.cell(2,3).value="Doctor"

sheet.cell(3,1).value=789
sheet.cell(3,2).value="Nicky"
sheet.cell(3,3).value="Teacher"

workbook.save(file)